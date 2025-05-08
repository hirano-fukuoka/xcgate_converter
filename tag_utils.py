from openpyxl.styles import PatternFill
import pandas as pd
from datetime import datetime, timedelta

def detect_tag_from_cell(cell):
    """背景色＋値ベースでタグを自動判定"""
    if not cell.fill or not isinstance(cell.fill, PatternFill):
        return "*入力 名前:'{}'".format(cell.value or "項目")
    
    rgb = cell.fill.fgColor.rgb if cell.fill.patternType == "solid" else None
    value = str(cell.value).lower() if cell.value else ""

    if rgb in ["FFFFFF00", "FFFF00"] or "日付" in value:
        return "*日付 名前:'点検日'"
    elif rgb == "FF00B0F0":
        return f"*数値 名前:'{cell.value}'"
    elif rgb == "FF00FF00":
        return f"*入力 名前:'{cell.value}'"
    elif rgb == "FFBFBFBF" or "実績" in value:
        return f"*実績 名前:'{cell.value}'"
    elif "選択" in value:
        return f"*選択 名前:'{cell.value}'"
    elif "送信" in value:
        return "*送信 名前:'実績送信'"
    else:
        return f"*入力 名前:'{cell.value}'"

def month_to_daily_df(ws):
    """1列目: 点検項目, 2列目以降: 各日 → DataFrameに変換"""
    item_names = [cell.value for cell in ws['A'] if cell.value]
    year = datetime.today().year
    month = datetime.today().month
    days = (datetime(year, month + 1, 1) - timedelta(days=1)).day

    data = []
    for day in range(1, days + 1):
        date_str = f"{year}/{month:02}/{day:02}"
        row = {"点検日": f"*日付 名前:'点検日' 初期値:'{date_str}'"}
        for item in item_names:
            row[item] = f"*入力 名前:'{item}'"
        data.append(row)

    return pd.DataFrame(data)
