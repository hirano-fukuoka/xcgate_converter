import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime, timedelta
from tag_utils import detect_tag_from_cell, month_to_daily_df

# タイトル
st.title("📋 Excel点検表 → XC-GATE帳票変換アプリ")

# サイドバーに取扱説明
st.sidebar.title("ℹ️ 取扱説明")
with st.sidebar.expander("▶️ アプリの使い方", expanded=True):
    st.markdown("""
### 📝 アプリ概要
このアプリは、**Excel点検表をXC-GATE帳票に自動変換**するツールです。

---

### ✅ 入力条件
- 点検表は `.xlsx` 形式
- 1列目：点検項目名
- 2列目以降：日別または月単位の点検データ

---

### 🎨 タグの変換ルール（背景色）
| 背景色 | 判定されるタグ |
|--------|----------------|
| 黄色   | `*日付`（点検日） |
| 青色   | `*数値`（数値入力） |
| 緑色   | `*入力`（文字入力） |
| グレー | `*実績`（表示項目） |

---

### 🧮 関数挿入（例）
- `=IF(B2="NG", "要対応", "")` → 自動判定
- `=NOW()` → 記録時間に使用

---

### 📤 出力
- 出力形式：`.xlsx`（XC-GATE帳票形式）
- 各日付1行、項目ごとにタグが付きます

---

### 🚀 よく使うタグ（自動で優先）
- `*入力`、`*数値`、`*実績`、`*選択`、`*送信`、`*日時`
""")

# メイン処理
uploaded_file = st.file_uploader("点検表Excelをアップロード", type=["xlsx"])
if uploaded_file:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    st.subheader("📄 元Excelプレビュー")
    df_raw = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows()])
    st.dataframe(df_raw)

    st.subheader("🗓️ 日次＋タグ付き帳票")
    df_daily = month_to_daily_df(ws)
    st.dataframe(df_daily)

    # Excel帳票の生成
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "XC-GATE帳票"

    for i, row in df_daily.iterrows():
        for j, val in enumerate(row):
            out_ws.cell(row=i+1, column=j+1, value=val)

    # ダウンロードボタン
    output = BytesIO()
    out_wb.save(output)
    st.download_button("📥 XC-GATE帳票をダウンロード", data=output.getvalue(), file_name="xcgate_output.xlsx")
